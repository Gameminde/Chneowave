#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Inspecte une feuille Excel et affiche les 3 premières lignes non vides
Usage:
  venv\Scripts\python importers\inspect_sheet.py --xls "base de donner/BDD INST 2025.xlsx" --sheet "Pl Metrologique 2025 "
"""
import argparse
from openpyxl import load_workbook

def main():
    p = argparse.ArgumentParser()
    p.add_argument('--xls', required=True)
    p.add_argument('--sheet', required=True)
    args = p.parse_args()

    wb = load_workbook(args.xls, data_only=True)
    if args.sheet not in wb.sheetnames:
        print(f"Feuille introuvable: {args.sheet}")
        return
    ws = wb[args.sheet]

    printed = 0
    for r in range(1, min(ws.max_row, 100) + 1):
        row_vals = ["" if c.value is None else str(c.value).strip() for c in ws[r]]
        if any(v for v in row_vals):
            printed += 1
            print(f"Ligne {r}: {row_vals}")
            if printed >= 3:
                break

if __name__ == '__main__':
    main()
