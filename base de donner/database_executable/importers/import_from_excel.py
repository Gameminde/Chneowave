#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Import des données métiers depuis BDD INST 2025.xlsx vers la base SQLite
- Source: chneowave/base de donner/BDD INST 2025.xlsx
- Cible:  instrumentation_maritime.db (SQLite)

Exige openpyxl (recommandé). Si indisponible, le script arrêtera avec un message clair.

Usage:
  python importers/import_from_excel.py --db instrumentation_maritime.db --xls "base de donner/BDD INST 2025.xlsx" --dry-run
  python importers/import_from_excel.py --db instrumentation_maritime.db --xls "base de donner/BDD INST 2025.xlsx"
"""

import argparse
import os
import sys
import sqlite3
from datetime import datetime
from typing import Dict, Any, List, Optional

try:
    from openpyxl import load_workbook
except Exception as e:
    print("ERREUR: Le module openpyxl est requis pour l'import depuis Excel.")
    print("Installez-le: pip install openpyxl")
    sys.exit(2)

SHEET_INV = "Inv 2025"
SHEET_METRO = "Pl Metrologique 2025 "
SHEET_PLAN_ETALO = "Suivi Plan Etalo 2025"
SHEET_LICENCES = "Licences "

# Dictionnaires de correspondance d'en-têtes possibles -> champ cible
EQUIPE_HEADERS = {
    "numero": ["numero", "n°", "n", "ref", "id", "code"],
    "description": ["description", "designation", "libelle", "libellé"],
    "marque_type": ["marque", "marque_type", "type", "modele", "modèle", "marque/type"],
    "numero_serie": ["serie", "n° serie", "n° série", "numero_serie", "sn", "s/n"],
    "n_inventaire": ["inventaire", "n_inventaire", "n° inventaire", "code inventaire", "n inv", "n_inv", "n inventaire", "n inv."],
    "utilisateur": ["utilisateur", "responsable", "affectation", "affecte"],
    "service": ["service", "srv", "dept", "departement", "département"],
    "localisation": ["localisation", "lieu", "emplacement", "site"],
    "etat": ["etat", "état", "status", "statut"],
    "annee_acquisition": ["annee", "année", "acquisition", "annee_acquisition"],
    "valeur_acquisition": ["valeur", "valeur_acquisition", "cout", "coût", "prix"],
    "statut_metrologique": ["statut metrologique", "statut_metro", "metrologie", "metrologique"],
    "date_derniere_verification": ["derniere verif", "dernière verif", "date derniere", "date dernière", "date_derniere_verification"],
    "prochaine_verification": ["prochaine verif", "prochaine verification", "echeance", "échéance", "prochaine_verification"],
    "frequence_verification_mois": ["frequence", "fréquence", "periode", "période"],
    "criticite": ["criticite", "criticité", "criticite_risque", "risque"],
    "commentaires": ["commentaires", "obs", "remarques", "notes"]
}

METRO_HEADERS = {
    "numero": ["numero", "n°", "code", "id_equipement", "id"],
    "n_inventaire": ["n inv", "n_inv", "n inventaire", "n° inventaire", "n inv.", "n inv "],
    "type_controle": ["type", "type_controle", "controle", "contrôle", "nature", "ext/int"],
    "date_verification": ["date", "date_verification", "date controle", "date contrôle", "date de verification", "date vérification"],
    "resultat": ["resultat", "résultat", "statut", "conformite", "conformité"],
    "prestataire": ["prestataire", "fournisseur", "laboratoire", "organisme"],
    "cout": ["cout", "coût", "montant"],
    "prochaine_echeance": ["prochaine", "echeance", "échéance", "prochaine_echeance"],
    "certificat_numero": ["certificat", "n cert", "n° cert", "ref certificat"]
}

SERVICE_CODE_MAP = {
    "CTS": "CTS",
    "CEM": "CEM",
    "DSC": "DSC",
    "INT": "INT"
}

def norm(s: Optional[str]) -> str:
    return (s or "").strip().lower().replace("\n", " ")

def find_columns(ws, header_map: Dict[str, List[str]]):
    """Retourne (mapping champ->index_col (1-based), header_row index)."""
    header_row = None
    for r in range(1, min(20, ws.max_row) + 1):
        values = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
        # heuristique: ligne avec >= 3 valeurs non vides
        if sum(1 for v in values if v) >= 3:
            header_row = r
            break
    if not header_row:
        return {}, None

    cols = {}
    values = [str(c.value).strip() if c.value is not None else "" for c in ws[header_row]]
    for idx, hv in enumerate(values, start=1):
        hvn = norm(hv)
        for field, aliases in header_map.items():
            if field in cols:
                continue
            if hvn in [norm(a) for a in aliases]:
                cols[field] = idx
    return cols, header_row

def get_or_create_service(conn, code: str, nom: Optional[str] = None) -> int:
    c = conn.cursor()
    c.execute("SELECT id FROM services WHERE code = ?", (code,))
    row = c.fetchone()
    if row:
        return row[0]
    c.execute("INSERT INTO services(code, nom) VALUES(?, ?)", (code, nom or code))
    conn.commit()
    return c.lastrowid

def parse_date(value) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, (datetime,)):
        return value.date().isoformat()
    if isinstance(value, str):
        v = value.strip()
        # Try common formats
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(v, fmt).date().isoformat()
            except Exception:
                pass
    return None

def parse_int(value) -> Optional[int]:
    try:
        if value is None or value == "":
            return None
        return int(str(value).strip().split(".")[0])
    except Exception:
        return None

def parse_float(value) -> Optional[float]:
    try:
        if value is None or value == "":
            return None
        v = str(value).replace(" ", "").replace(",", ".")
        return float(v)
    except Exception:
        return None

def import_equipements(conn, ws) -> int:
    result = find_columns(ws, EQUIPE_HEADERS)
    cols, header_row = (result if isinstance(result, tuple) else (result, 1))
    if not cols:
        print(f"[EQUIPEMENTS] En-têtes non trouvés sur la feuille '{ws.title}'. Abandon.")
        return 0

    inserted = 0
    cur = conn.cursor()

    for r in range((header_row or 1) + 1, ws.max_row + 1):
        def cell(colname):
            idx = cols.get(colname)
            return ws.cell(row=r, column=idx).value if idx else None

        numero = (cell("numero") or "").strip()
        n_inv = (cell("n_inventaire") or "").strip()
        description = (cell("description") or "").strip()
        if not numero and not n_inv and not description:
            continue

        service_raw = (cell("service") or "").strip().upper()
        service_code = SERVICE_CODE_MAP.get(service_raw, service_raw[:3] if service_raw else None)
        service_id = get_or_create_service(conn, service_code, service_code) if service_code else None

        # Prépare les données
        # Construire un numero stable: si pas de numero mais inventaire présent, utilise INV-<n_inv>
        numero_stable = numero or (f"INV-{n_inv}" if n_inv else f"EQ-{r:03d}")
        data = {
            "numero": numero_stable,
            "description": description or "(sans description)",
            "marque_type": (cell("marque_type") or "").strip(),
            "numero_serie": (cell("numero_serie") or "").strip(),
            "n_inventaire": n_inv or None,
            "utilisateur": (cell("utilisateur") or "").strip() or None,
            "service_id": service_id,
            "localisation": (cell("localisation") or "").strip() or None,
            "etat": (cell("etat") or "OK").strip() or "OK",
            "annee_acquisition": parse_int(cell("annee_acquisition")),
            "valeur_acquisition": parse_float(cell("valeur_acquisition")),
            "statut_metrologique": (cell("statut_metrologique") or "CONFORME").strip() or "CONFORME",
            "date_derniere_verification": parse_date(cell("date_derniere_verification")),
            "prochaine_verification": parse_date(cell("prochaine_verification")),
            "frequence_verification_mois": parse_int(cell("frequence_verification_mois")) or 12,
            "criticite": parse_int(cell("criticite")) or 1,
            "commentaires": (cell("commentaires") or "").strip() or None,
        }

        # Recherche existants: priorité à l'inventaire, sinon par numero_stable
        row_inv = None
        if n_inv:
            cur.execute("SELECT id_equipement FROM equipements WHERE n_inventaire = ?", (n_inv,))
            row_inv = cur.fetchone()
        row_num = None
        if not row_inv:
            cur.execute("SELECT id_equipement FROM equipements WHERE numero = ?", (numero_stable,))
            row_num = cur.fetchone()

        target_id = row_inv[0] if row_inv else (row_num[0] if row_num else None)

        if target_id:
            # Eviter conflit unique sur n_inventaire
            data_update = dict(data)
            if n_inv:
                cur.execute("SELECT id_equipement FROM equipements WHERE n_inventaire = ?", (n_inv,))
                row_check = cur.fetchone()
                if row_check and row_check[0] != target_id:
                    data_update.pop("n_inventaire", None)
            else:
                data_update.pop("n_inventaire", None)

            placeholders = ", ".join([f"{k} = ?" for k in data_update.keys() if k != "numero"])  # on évite de changer numero clé
            values = [v for k, v in data_update.items() if k != "numero"]
            values.append(target_id)
            cur.execute(f"UPDATE equipements SET {placeholders} WHERE id_equipement = ?", values)
        else:
            # Tentative d'insertion, fallback en UPDATE si contrainte d'unicité
            cols_list = ",".join(data.keys())
            qs = ",".join(["?"] * len(data))
            try:
                cur.execute(f"INSERT INTO equipements ({cols_list}) VALUES ({qs})", list(data.values()))
                inserted += 1
            except sqlite3.IntegrityError:
                # Conflit sur numero ou n_inventaire: faire un UPDATE par numero_stable
                placeholders = ", ".join([f"{k} = ?" for k in data.keys() if k != "numero"])  # ne pas changer numero
                values = [v for k, v in data.items() if k != "numero"]
                values.append(numero_stable)
                cur.execute(f"UPDATE equipements SET {placeholders} WHERE numero = ?", values)

    conn.commit()
    print(f"[EQUIPEMENTS] Lignes insérées/maj: {inserted}")
    return inserted

def import_metrologie(conn, ws) -> int:
    result = find_columns(ws, METRO_HEADERS)
    cols, header_row = (result if isinstance(result, tuple) else (result, 1))
    if not cols:
        print(f"[METROLOGIE] En-têtes non trouvés sur la feuille '{ws.title}'. Abandon.")
        return 0
    cur = conn.cursor()
    inserted = 0

    for r in range((header_row or 1) + 1, ws.max_row + 1):
        def cell(colname):
            idx = cols.get(colname)
            return ws.cell(row=r, column=idx).value if idx else None

        numero = (cell("numero") or "").strip()
        inv = (cell("n_inventaire") or "").strip()
        key = numero or inv
        if not key:
            continue
        # map numero/inventaire -> id_equipement
        cur.execute("SELECT id_equipement FROM equipements WHERE numero = ?", (key,))
        row = cur.fetchone()
        if not row:
            cur.execute("SELECT id_equipement FROM equipements WHERE n_inventaire = ?", (key,))
            row = cur.fetchone()
        if not row:
            continue
        equip_id = row[0]

        ext_int = (cell("type_controle") or "").strip().upper()
        # Si la colonne détectée est EXT/INT, déduire type/prestataire
        if ext_int in ("EXT", "INTERNE", "INT"):
            type_controle = "ETALONNAGE" if ext_int == "EXT" else "VERIFICATION"
            prestataire = "EXTERNE" if ext_int == "EXT" else "INTERNE"
        else:
            type_controle = ext_int or "ETALONNAGE"
            prestataire = (cell("prestataire") or "").strip() or None

        data = {
            "equipement_id": equip_id,
            "type_controle": type_controle,
            "date_verification": parse_date(cell("date_verification")) or datetime.now().date().isoformat(),
            "resultat": (cell("resultat") or "CONFORME").strip(),
            "prestataire": prestataire,
            "cout": parse_float(cell("cout")),
            "devise": "DA",
            "prochaine_echeance": parse_date(cell("prochaine_echeance")),
            "certificat_numero": (cell("certificat_numero") or "").strip() or None,
            "certificat_path": None,
            "commentaires": None,
            "technicien_responsable": None,
        }
        cols_list = ",".join(data.keys())
        qs = ",".join(["?"] * len(data))
        cur.execute(f"INSERT INTO metrologie ({cols_list}) VALUES ({qs})", list(data.values()))
        inserted += 1

    conn.commit()
    print(f"[METROLOGIE] Lignes insérées: {inserted}")
    return inserted


def main():
    parser = argparse.ArgumentParser(description="Import Excel vers SQLite (Instrumentation Maritime)")
    parser.add_argument("--db", default="instrumentation_maritime.db", help="Chemin du fichier SQLite")
    parser.add_argument("--xls", default=os.path.join("base de donner", "BDD INST 2025.xlsx"), help="Chemin du fichier Excel")
    parser.add_argument("--dry-run", action="store_true", help="Ne pas écrire dans la base, juste simuler")
    args = parser.parse_args()

    if not os.path.exists(args.xls):
        print(f"ERREUR: Fichier Excel introuvable: {args.xls}")
        sys.exit(1)
    if not os.path.exists(args.db):
        print(f"ERREUR: Base SQLite introuvable: {args.db}. Créez-la d'abord (create_db_simple.py)")
        sys.exit(1)

    wb = load_workbook(args.xls, data_only=True)

    # Connexion DB
    conn = sqlite3.connect(args.db)
    conn.row_factory = sqlite3.Row

    total_eq = total_metro = 0

    if SHEET_INV in wb.sheetnames:
        ws_inv = wb[SHEET_INV]
        if args.dry_run:
            print(f"[DRY-RUN] Analyse '{SHEET_INV}' - lignes: {ws_inv.max_row}")
        else:
            total_eq = import_equipements(conn, ws_inv)
    else:
        print(f"AVERTISSEMENT: Feuille '{SHEET_INV}' absente")

    if SHEET_METRO in wb.sheetnames:
        ws_m = wb[SHEET_METRO]
        if args.dry_run:
            print(f"[DRY-RUN] Analyse '{SHEET_METRO}' - lignes: {ws_m.max_row}")
        else:
            total_metro = import_metrologie(conn, ws_m)
    else:
        print(f"AVERTISSEMENT: Feuille '{SHEET_METRO}' absente")

    print("\nRÉSUMÉ IMPORT:")
    print(f"  Equipements insérés/maj: {total_eq}")
    print(f"  Enregistrements métrologie: {total_metro}")

    conn.close()

if __name__ == "__main__":
    main()
